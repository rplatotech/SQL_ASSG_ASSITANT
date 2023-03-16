import openpyxl
import expected_lists_for_questions
import pymysql.cursors
import decimal
import datetime
from gen_results_from_excel_queries import sql_logs_reader, sql_logs_clear


def sql_things(query):
    global rows

    # Define a converter function to convert pymysql data types to Python built-in types
    def convert_pymysql_types(data):
        if isinstance(data, (bytearray, bytes)):
            return data.decode('utf-8')
        elif isinstance(data, (int, float)):
            return data
        elif isinstance(data, decimal.Decimal):
            return float(data)
        elif isinstance(data, (datetime.date, datetime.datetime)):
            return data.isoformat()
        elif data is None:
            return None
        else:
            return str(data)

    # # Load the Excel sheet and extract the SQL statements
    # wb = openpyxl.load_workbook('query_excel_file.xlsx')
    # ws = wb.active
    # queries = [cell.value for cell in ws['A'] if cell.value.startswith('SELECT')]

    # Define the database connection parameters
    host = '127.0.0.1'
    user = 'root'
    password = 'password'
    database = 'sql_invoicing'

    # Connect to the database
    conn = pymysql.connect(host=host,
                           user=user,
                           password=password,
                           db=database,
                           # cursorclass=pymysql.cursors.DictCursor)
                           )

    # Execute a SELECT statement
    # for query in queries:

    with conn.cursor() as cur:
        cur.execute(query)
        results = cur.fetchall()

    # Convert the data types in the query results and store in a list of lists
    rows = []
    for result in results:
        row = []
        for col in result:
            row.append(str(convert_pymysql_types(col)))
        rows.append(row)

    return rows
    # Get the column names from the keys of the first dictionary in the query results
    col_names = list(results)


# Print the query results
# print(col_names)
def read_query_from_excel(cell):
    wb = openpyxl.load_workbook('query_excel_file.xlsx')
    ws = wb['Sheet1']
    value = ws[cell].value
    return value


def test_a2_q1():
    sql_things(query=read_query_from_excel('A1'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q1, 'Query error. Check manually'

def test_a2_q2():
    sql_things(query=read_query_from_excel('A2'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q2, 'Query error. Check manually'

def test_a2_q3():
    sql_things(query=read_query_from_excel('A3'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q3, 'Query error. Check manually'

def test_a2_q4():
    sql_things(query=read_query_from_excel('A4'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q4, 'Query error. Check manually'


def test_a2_q5():
    sql_things(query=read_query_from_excel('A5'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q5, 'Query error. Check manually'

def test_a2_q6():
    sql_things(query=read_query_from_excel('A6'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q6, 'Query error. Check manually'


def test_a2_q7():
    sql_things(query=read_query_from_excel('A7'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q7, 'Query error. Check manually'

def test_a2_q8():
    sql_things(query=read_query_from_excel('A8'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q8, 'Query error. Check manually'

def test_a2_q9():
    sql_things(query=read_query_from_excel('A9'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q9, 'Query error. Check manually'

def test_a2_q10():
    sql_things(query=read_query_from_excel('A10'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q10, 'Query error. Check manually'

def test_a2_q11():
    sql_things(query=read_query_from_excel('A11'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q11, 'Query error. Check manually'

def test_a2_q12():
    sql_things(query=read_query_from_excel('A12'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q12, 'Query error. Check manually'

def test_a2_q13():
    sql_things(query=read_query_from_excel('A13'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q13, 'Query error. Check manually'

def test_a2_q14():
    sql_things(query=read_query_from_excel('A14'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q14, 'Query error. Check manually'


def test_a2_q15():
    sql_things(query=read_query_from_excel('A15'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el_q15, 'Query error. Check manually'

# ASSIGNMENT 3 ---------------------------------------------

def test_a3_q1():
    sql_things(query=read_query_from_excel('A16'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q1, 'Query error. Check manually'

def test_a3_q2():
    sql_things(query=read_query_from_excel('A17'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q2, 'Query error. Check manually'

def test_a3_q3():
    sql_things(query=read_query_from_excel('A18'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q3, 'Query error. Check manually'

def test_a3_q4():
    sql_things(query=read_query_from_excel('A19'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q4, 'Query error. Check manually'

def test_a3_q5():
    sql_things(query=read_query_from_excel('A20'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q5, 'Query error. Check manually'

def test_a3_q6():
    sql_things(query=read_query_from_excel('A21'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q6, 'Query error. Check manually'

def test_a3_q7():
    sql_things(query=read_query_from_excel('A22'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q7, 'Query error. Check manually'

def test_a3_q8():
    sql_things(query=read_query_from_excel('A23'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q8, 'Query error. Check manually'

def test_a3_q9():
    sql_things(query=read_query_from_excel('A24'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q9, 'Query error. Check manually'

def test_a3_q10():
    sql_things(query=read_query_from_excel('A25'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el3_q10, 'Query error. Check manually'
#
# # ASSIGNMENT 4 --------------------------------------

def test_a4_q1():
    sql_things(query=read_query_from_excel('A26'))
    # print(read_query_from_excel('A26'))
    print('\n')
    for row in rows:
        print(row)
        print(sql_logs_reader())
        assert sql_logs_reader() in expected_lists_for_questions.el4_q1
    sql_logs_clear()

def test_a4_q2():
    sql_things(query=read_query_from_excel('A27'))
    # print(read_query_from_excel('A26'))
    print('\n')
    for row in rows:
        print(row)
        print(sql_logs_reader())
        assert sql_logs_reader() in expected_lists_for_questions.el4_q2
    sql_logs_clear()

def test_a4_q3():
    sql_things(query=read_query_from_excel('A28'))
    # print(read_query_from_excel('A26'))
    print('\n')
    for row in rows:
        print(row)
        print(sql_logs_reader())
        assert sql_logs_reader() in expected_lists_for_questions.el4_q3
    # sql_logs_clear()

def test_a4_q4():
    sql_things(query=read_query_from_excel('A29'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q4, 'Query error. Check manually'

def test_a4_q5():
    sql_things(query=read_query_from_excel('A30'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q5, 'Query error. Check manually'

def test_a4_q6():
    sql_things(query=read_query_from_excel('A31'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q6, 'Query error. Check manually'

def test_a4_q7():
    sql_things(query=read_query_from_excel('A32'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q7, 'Query error. Check manually'

def test_a4_q8():
    sql_things(query=read_query_from_excel('A33'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q8, 'Query error. Check manually'

def test_a4_q9():
    sql_things(query=read_query_from_excel('A34'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q9, 'Query error. Check manually'

def test_a4_q10():
    sql_things(query=read_query_from_excel('A35'))
    print('\n')
    for row in rows:
        print(row)
        assert row in expected_lists_for_questions.el4_q10, 'Query error. Check manually'
